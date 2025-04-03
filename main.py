import customtkinter as ctk
from constants import *
from PIL import Image, ImageTk
from convert_pdf import PDFFile
from grid import Grid
import cv2
from processing import process_pdf
import openpyxl

# TODO: make warning class
# TODO: Add Errors
# TODO: Show the number of files


class ImageViewer(ctk.CTkCanvas):
    def __init__(self, master, cv_image):
        super().__init__(master=master, bg="#1F1F1F")  # Set background in constructor

        if App.cv_image is None:
            App.cv_image = cv_image

        self.grid(column=0, columnspan=2, row=0, sticky="nsew")
        self.bind("<Configure>", self.resize_callback)

        # Zoom and Pan Variables
        self.scale = 1.0
        self.offset_x = 0
        self.offset_y = 0
        self.last_x = 0
        self.last_y = 0

        # Bind Mouse Events
        self.bind("<MouseWheel>", self.zoom)  # Windows & Mac
        self.bind("<ButtonPress-1>", self.start_pan)  # Left-click to start panning
        self.bind("<B1-Motion>", self.pan)  # Drag to pan
        self.bind("<ButtonRelease-1>", self.end_pan)

    def resize(self, width, height):
        if App.cv_image is None:
            return

        # Convert BGR (OpenCV) to RGB (PIL)
        b, g, r = cv2.split(App.cv_image)
        re_image = cv2.merge((r, g, b))
        image = Image.fromarray(re_image)

        # Apply Zoom
        new_size = (int(image.width * self.scale), int(image.height * self.scale))
        image = image.resize(new_size, Image.LANCZOS)

        # Store as Tkinter-compatible image
        App.resized_tk_img = ImageTk.PhotoImage(image)

    def resize_callback(self, event):
        if App.cv_image is None:
            return

        width, height = self.winfo_width(), self.winfo_height()
        self.delete("all")
        self.resize(width, height)
        self.redraw_image()

    def redraw_image(self):
        """Redraws the image at the correct position with current zoom and pan."""
        if App.resized_tk_img:
            self.delete("all")
            self.create_image(
                self.winfo_width() // 2 + self.offset_x,
                self.winfo_height() // 2 + self.offset_y,
                image=App.resized_tk_img,
                anchor="center",
            )

    def zoom(self, event):
        """Zooms in/out using the mouse wheel."""
        zoom_factor = 1.1 if event.delta > 0 else 0.9
        new_scale = self.scale * zoom_factor

        if 0 < new_scale < 3:
            self.scale = new_scale
            self.resize(self.winfo_width(), self.winfo_height())
            self.redraw_image()

    def start_pan(self, event):
        """Records the starting point of a pan."""
        self.last_x = event.x
        self.last_y = event.y

    def pan(self, event):
        """Moves the image based on mouse movement."""
        dx = event.x - self.last_x
        dy = event.y - self.last_y

        self.offset_x += dx
        self.offset_y += dy
        self.redraw_image()

        self.last_x = event.x
        self.last_y = event.y

    def end_pan(self, event):
        """Stops panning."""
        pass  # Not needed, but included for future expansion


class NavigationButtons(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master=master)
        self.grid(row=1, columnspan=2, column=0, sticky="ew", padx=7, pady=4)

        self.columnconfigure(0, weight=1, uniform="b")
        self.columnconfigure(1, weight=1, uniform="b")
        self.columnconfigure(2, weight=1, uniform="b")
        self.columnconfigure(3, weight=1, uniform="b")
        self.columnconfigure(4, weight=1, uniform="b")

        self.previous_file = ctk.CTkButton(self, text="<<")
        self.previous_file.grid(column=0, row=0, padx=10, pady=15)

        self.next_file = ctk.CTkButton(self, text=">>")
        self.next_file.grid(column=4, row=0, padx=20, pady=15)

        self.previous_image = ctk.CTkButton(self, text="<")
        self.previous_image.grid(column=1, row=0, padx=10, pady=15)

        self.next_image = ctk.CTkButton(self, text=">")
        self.next_image.grid(column=3, row=0, padx=10, pady=15)

    def set_command_functions(self, functions):
        self.previous_file.configure(command=functions[0])
        self.previous_image.configure(command=functions[1])
        self.next_image.configure(command=functions[2])
        self.next_file.configure(command=functions[3])


class App:
    resized_tk_img = None
    cv_image = None
    decoded_info = None
    hover = None

    def __init__(
        self,
        title: str,
        width: int,
        height: int,
        resizable: list[bool],
        icon_path: str = None,
    ):
        self.width = width
        self.height = height
        self.pdfs: list[PDFFile] = []
        self.current_pdf: int | None = None
        self.current_grid: Grid | None = None

        self.window = ctk.CTk()
        self.window.geometry(f"{self.width}x{self.height}")
        self.window.resizable(resizable[0], resizable[1])
        self.window.title(title)

        self.window.columnconfigure((0, 1, 2, 3, 4), weight=1, uniform="a")
        self.window.rowconfigure(0, weight=1)
        self.window.rowconfigure(1, weight=0)

        self.rhs_frame = ctk.CTkFrame(self.window)
        self.rhs_frame.grid(
            row=0, rowspan=2, column=2, columnspan=4, padx=4, pady=10, sticky="nsew"
        )

        self.rhs_frame.rowconfigure((0, 1, 2, 3, 4), weight=1, uniform="a")
        self.rhs_frame.columnconfigure((0, 1), weight=1, uniform="a")

        # Define StringVar variables
        self.grid_type = ctk.StringVar(value="Type de Grille: N/A")
        self.score = ctk.StringVar(value="Score : N/A")

        self.grid_type_label = ctk.CTkLabel(self.rhs_frame, textvariable=self.grid_type)

        self.score_label = ctk.CTkLabel(
            self.rhs_frame, text="Score: ", textvariable=self.score
        )

        self.export_button = ctk.CTkButton(
            self.rhs_frame, text="Exporter", command=self.save_file
        )
        self.add_file_button = ctk.CTkButton(
            self.rhs_frame, text="Ajouter", command=self.open_file
        )

        self.grid_type_label.grid(
            row=0, column=0, columnspan=2, padx=10, pady=5, sticky="nsew"
        )
        self.score_label.grid(
            row=3, column=0, padx=10, columnspan=2, pady=5, sticky="nsew"
        )
        self.export_button.grid(
            row=4, column=0, columnspan=1, padx=10, pady=10, sticky="ew"
        )
        self.add_file_button.grid(
            row=4, column=1, columnspan=1, padx=10, pady=10, sticky="ew"
        )

        self.rhs_frame.grid_columnconfigure(0, weight=1)

        self.bottom_bottons = NavigationButtons(self.window)
        self.bottom_bottons.set_command_functions(
            [
                self.show_previous_pdf,
                self.show_previous_grid,
                self.show_next_grid,
                self.show_next_pdf,
            ]
        )

        if icon_path:
            self.window.iconbitmap(icon_path)

        self.image_viewer = ImageViewer(self.window, App.cv_image)

    def update(self, data: tuple = None):
        if data is None:
            return
        App.cv_image = data["image"]
        self.grid_type.set(value=data["type"].value)
        self.image_viewer.after(
            10, lambda: self.image_viewer.event_generate("<Configure>")
        )

        data = self.current_grid.find_multiple_checks_and_empty_rows()

        if hasattr(self, "selection_frame"):
            self.selection_frame.destroy()
        self.show_selection_frame(data)
        if not data["empty_rows"] and not data["multiple_detections"]:
            self.score.set(value=f"Score:{self.current_grid.calculate_score()}")

    def _set_current_grid(self, event=None):
        self.update(self.current_grid.run_analysis())

    def _show_latest_file(self, event=None):
        if self.current_pdf is None:
            self.current_pdf = 0
        else:
            self.current_pdf = len(self.pdfs) - 1
        self.pdfs[self.current_pdf].extract_grids()
        self.current_grid = self.pdfs[self.current_pdf].get_next_grid()
        self._set_current_grid()

    def show_next_pdf(self, event=None):
        if not self.pdfs or self.current_pdf is None:
            return
        self.current_pdf = (self.current_pdf + 1) % len(self.pdfs)
        self.pdfs[self.current_pdf].extract_grids()
        self.current_grid = self.pdfs[self.current_pdf].get_next_grid()
        self._set_current_grid()

    def show_previous_pdf(self, event=None):
        if not self.pdfs or self.current_pdf is None:
            return
        self.current_pdf = (self.current_pdf + 1) % len(self.pdfs)
        self.pdfs[self.current_pdf].extract_grids()
        self.current_grid = self.pdfs[self.current_pdf].get_previous_grid()
        self._set_current_grid()

    def show_next_grid(self, event=None):
        if not self.pdfs or self.current_pdf is None:
            return
        self.current_grid = self.pdfs[self.current_pdf].get_next_grid()
        self._set_current_grid()

    def show_previous_grid(self, event=None):
        if not self.pdfs or self.current_pdf is None:
            return
        self.current_grid = self.pdfs[self.current_pdf].get_previous_grid()
        self._set_current_grid()

    def open_file(self, event=None):
        file_paths = ctk.filedialog.askopenfilenames(
            title="Sélectionner un fichier", filetypes=[("Fichiers PDF", "*.pdf")]
        )
        for file_path in file_paths:
            self.pdfs.append(PDFFile(file_path))
        self._show_latest_file()

    def save_file(self, event=None):
        if self.current_grid is None:
            return

        file_path = ctk.filedialog.asksaveasfilename(
            title="Sauvegarder fichier excel",
            filetypes=[("Fichiers Excel", "*.xlsx")],
            defaultextension=".xlsx",
        )
        if not file_path:
            return

        xl_original_grid_obj = openpyxl.load_workbook("documents/Grille.xlsx")
        xl_original_grid_obj.save(file_path)
        xl_original_grid_obj.close()

        checked_cells = self.current_grid.cells_state
        xl_copy = openpyxl.load_workbook(file_path)
        active_sheet = xl_copy.active
        for i, row in enumerate(checked_cells, start=3):
            for j, cell in enumerate(row, start=4):
                active_sheet.cell(row=i, column=j).value = "X" if cell[0] > 0 else ""

        xl_copy.save(file_path)
        xl_copy.close()

    def run(self):
        self.update()
        self.window.mainloop()

    # TODO: make warning class
    # TODO: Add Errors
    # TODO: Show the number of files

    def show_selection_frame(self, warnings_dict):
        """Creates a selection frame on the right half of the window for user input."""

        # Create a new frame inside rhs_frame (right side)
        self.selection_frame = ctk.CTkFrame(self.rhs_frame)
        self.selection_frame.grid(
            row=1,rowspan=2, column=0, columnspan=2, padx=10, pady=10, sticky="nsew"
        )

        # Title label
        label = ctk.CTkLabel(
            self.selection_frame, text="⚠ Select One Cell Per Row", font=("Arial", 14)
        )
        label.pack(pady=5)

        # Handle multiple checks (rows with multiple checked cells)
        multiple_checks = warnings_dict.get("multiple_detections", [])
        for row_index, checked_cells in multiple_checks:
            row_frame = ctk.CTkFrame(self.selection_frame)
            row_frame.pack(pady=5, fill="x")

            row_label = ctk.CTkLabel(
                row_frame, text=f"Row {row_index + 1}: ", font=("Arial", 12)
            )
            row_label.pack(side="left", padx=5)

            for col_index in checked_cells:
                btn = ctk.CTkButton(
                    row_frame,
                    text=f"Cell {col_index + 1}",
                    command=lambda r=row_index, c=col_index: self.select_cell(r, c),
                    fg_color="blue",
                    hover_color="darkblue",
                )
                btn.pack(side="left", padx=5)

                # Bind hover events to highlight the corresponding cell in the OpenCV image
                btn.bind(
                    "<Enter>",
                    lambda event, r=row_index, c=col_index: self.on_hover(r, c),
                )
                btn.bind(
                    "<Leave>",
                    lambda event, r=row_index, c=col_index: self.on_leave(r, c),
                )

        # Display other warnings
        other_warnings = warnings_dict.get("other_warnings", [])
        if other_warnings:
            warnings_frame = ctk.CTkFrame(self.selection_frame)
            warnings_frame.pack(pady=10, fill="x")

            warnings_label = ctk.CTkLabel(
                warnings_frame, text="⚠ Warnings", font=("Arial", 14, "bold")
            )
            warnings_label.pack(pady=5)

            # Display each warning in the other_warnings list
            for warning in other_warnings:
                warning_label = ctk.CTkLabel(
                    warnings_frame, text=warning, font=("Arial", 12)
                )
                warning_label.pack(anchor="w", padx=10, pady=3)

        # Display empty rows (rows that are empty)
        empty_rows = warnings_dict.get("empty_rows", [])
        if empty_rows:
            empty_rows_frame = ctk.CTkFrame(self.selection_frame)
            empty_rows_frame.pack(pady=10, fill="x")

            empty_rows_label = ctk.CTkLabel(
                empty_rows_frame, text="❌ Empty Rows", font=("Arial", 14, "bold")
            )
            empty_rows_label.pack(pady=5)

            # Display each empty row
            for empty_row in empty_rows:
                empty_row_label = ctk.CTkLabel(
                    empty_rows_frame, text=empty_row, font=("Arial", 12)
                )
                empty_row_label.pack(anchor="w", padx=10, pady=3)

        # Make sure the right side resizes properly
        self.rhs_frame.rowconfigure(1, weight=1)

    def on_hover(self, row, col):
        """Highlights the corresponding cell in the OpenCV image when hovering over an option."""
        if self.current_grid and App.hover != (row, col):
            self.current_grid.change_cell_color(row, col, 2)
            App.cv_image = self.current_grid.drawn_og_img
            self.image_viewer.after(
                10, lambda: self.image_viewer.event_generate("<Configure>")
            )
            App.hover = (row, col)

    def on_leave(self, row, col):
        """Removes the highlight when the user moves the mouse away."""
        # self.image_viewer.update_image(App.cv_image)
        if self.current_grid:
            self.current_grid.change_cell_color(row, col, 3)
            App.cv_image = self.current_grid.drawn_og_img
            self.image_viewer.after(
                10, lambda: self.image_viewer.event_generate("<Configure>")
            )
            App.hover = None

    def select_cell(self, row, col):
        """Updates the grid state to keep only the selected cell checked and manages warnings."""
        self.current_grid.set_selected_cell(row, col)

        # Re-check for warnings
        data = self.current_grid.find_multiple_checks_and_empty_rows()
        self.show_selection_frame(data)
        # If no more warnings, remove the selection frame
        if not data['empty_rows'] and not data['multiple_detections']:
            self.selection_frame.destroy()
            del self.selection_frame
            self.score.set(value=f"Score:{self.current_grid.calculate_score()}")


app = App("Demo", WIDTH, HEIGHT, [True, True])
app.run()
